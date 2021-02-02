# -*- coding: utf-8 -*-
# !/usr/bin/env python3

import pandas as pd
import datetime
import email
import smtplib
import os
import openpyxl
from openpyxl import load_workbook
import os
x = os.listdir()  # список файлов в текущей папке
for i in x:  # ищем среди файлов начинающееся с "Отчет"
    if i[0:5] == 'Отчет':  #
        name_file = i  # присваивается название файла переменной name_file

        new_file_name_1 = '1 ' + name_file  # добавит 1 к назв.файла, если есть 'Салон'
        new_file_name_2 = '2 ' + name_file  # добавит 2 к назв.файла, если есть нету 'Салон'

        wb = load_workbook(filename=name_file)  # вставляем файл-эксель и присваеваем к wb
        wb.active = 0  # активируем первый sheetname (нулевой)
        sheet = wb.active  # присваиваем работу с листом к переменной sheet

        if sheet.title == 'Лист1':  # если Лист1, то меняем название на 'Период'
            sheet.title = 'Период'
            sheet['A2'].value = 'period'  # и в ячейку A2 вставляем 'period'

        # переходим к листу 'По дням'
        wb.active = 1  # активируем второй лист.
        sheet2 = wb.active  # присваиваем работу с листом к переменной sheet2

        if sheet2.title == 'По дням':  # если лист точно 'По дням', то вставляем в A2 'day'
            sheet2['A2'].value = 'day'

        # проставляем даты
        all_data = []  # тут будут все даты
        first_data = []  # тут будет первая дата
        last_data = []  # тут будет последняя дата

        data = [sheet2.cell(row=4, column=i).value for i in
                range(1, 1000)]  # проходимся по ряду где дата,проверим 1000 ячеек,чтобы навярняка охватить

        for i in data:  # добавляем все даты и избавляемся от None
            if i != None:
                all_data.append(i)
        first_data = [all_data[0]]  # Добавляем первую дату
        last_data = [all_data[-1]]  # Добаляем последнюю дату

        # преобразовываем формат даты yyyy/mm/dd в dd.mm.yyyy
        yyyy, mm, dd = first_data[0].split('/')  # разбираем на отдельные части yyyy, mm,dd
        first_data[0] = '{}-{}-{}'.format(dd[0:2], mm, yyyy)  # готовим минимальную дату к нужному формату
        sheet['B2'].value = first_data[0]  # вставляем в ячейку B2 минимальную дату

        yyyy, mm, dd = last_data[0].split('/')  # разбираем на отдельные части yyyy, mm,dd
        last_data[0] = '{}-{}-{}'.format(dd[0:2], mm, yyyy)  # готовим максимальную дату к нужному формату
        sheet['C2'].value = last_data[0]  # вставляем в ячейку С2 максимальную дату

        # сохраняем файл с некоторыми условиями
        data2 = [sheet.cell(row=i, column=3).value for i in range(1, 30)]  # набираем данные из ячеек
        data3 = [x for x in data2 if x is not None]  # убираем из списка None

        for i in data3:  # проверяем есть ли слово "Салон", и сохраняем с определенным образом.
            if 'Салон' in str(i):
                wb.save(new_file_name_1)
                break
        else:
            wb.save(new_file_name_2)


def report1_daily(file, sheet):
    df_input = pd.read_excel(file,sheet_name=sheet, skiprows=3, )
    columns = df_input.columns
#    ----------------------------------------------------------
    dates = []
    for col in columns:
        if col.startswith('Unnamed'):
            continue
        else:
            dates.append(col)
#--------------------------------------------------------------
    for col in df_input.columns[4:]:
            if col.startswith('Unnamed'):
                df_input.rename(columns={col: last_date}, inplace=True)
            else:
                last_date = col
#    ----------------------------------------------------------
    banks = list(set(df_input.loc[0]))[1:]
#    ----------------------------------------------------------
    df_input.rename(columns={'Unnamed: 0': 'Дирекция', 'Unnamed: 1': 'Город', 'Unnamed: 2':'Магазин', 'Unnamed: 3': 'Тип продажи'}, inplace=True)
    df_input.fillna('None', inplace=True)
#    ----------------------------------------------------------
    channel = "РБТ"
    output_df2 = []
    last_direction = df_input.loc[1, 'Дирекция']
    last_town = df_input.loc[1, 'Город']
    for i in range(1, df_input.shape[0]):
        if df_input.loc[i, 'Дирекция'] == 'Местонахождение':
            break
        if df_input.loc[i, 'Дирекция'] != 'None':
                last_direction = df_input.loc[i, 'Дирекция']
        if df_input.loc[i, 'Город'] != 'None':
                last_town = df_input.loc[i, 'Город']
        for j in range(4,len(df_input.columns[4:]) + 4):
            if df_input.iloc[i, j] == 0:
                continue
            output_list = []
            output_list.append('d') #d_type
            output_list.append(channel) #channel
            output_list.append(df_input.loc[i, 'Магазин']) #sellerplace_group
            output_list.append(datetime.datetime.strptime(df_input.columns[j].strip(), '%Y/%m/%d')) #date_
            output_list.append(df_input.iloc[0, j]) #competitor
            output_list.append(df_input.iloc[i, j]) #credit_amt
            output_list.append(datetime.date.today()) #date_update
            output_df2.append(output_list)
#    ----------------------------------------------------------
    out_to_excel = pd.DataFrame(data=output_df2, columns=['D_TYPE', 'CHANNEL', 'SELLERPLACE_GROUP', 'DATE_', 'COMPETITOR', 'CREDIT_AMT', 'DATE_UPDATE'])
    out_to_excel.to_excel('new_dataForm_report1_daily.xlsx', index=False)
    return list(out_to_excel.values.tolist()), out_to_excel.columns

df = ['LDLDLD', 'DLDLDLD', 'lDmPds']
df_lower = [x.lower() for x in df]
df_lower

def report1_period(file, sheet, period_start, period_end):
    df_input = pd.read_excel(file,sheet_name=sheet, skiprows=2) # Вставляем Excel файл с данными
    columns_list = []
    df_input.rename(columns={'Показатели\nas values':'Дирекция', 'Unnamed: 1':'Город', 'Unnamed: 2':'Магазин', 'Unnamed: 3':'Тип продажи'}, inplace=True)
    columns_lower = [column.lower() for column in list(df_input.columns)]
    for column in df_input.columns[columns_lower.index('итого'):columns_lower.index('итого') + 3]:
        columns_list.append(df_input.loc[0, column])
        df_input.rename(columns={column: df_input.loc[0, column]}, inplace=True)
#    ----------------------------------------------------------
#     bank_list = []
#     for i in df_input.columns[4:actual_columns(df_input)-2]:
#         if i.startswith('Unnamed'):
#             continue
#         else:
#             bank_list.append(i)
# #    ----------------------------------------------------------
#     for column in df_input.columns[4: actual_columns(df_input)]:
#         new_column_name = column + ' ' + df_input.loc[0, column][:-1]
#         df_input.rename(columns={column: new_column_name}, inplace=True)

#     for column in df_input.columns[4:]:
#         if column.endswith('Продаж в кредит'):
#             bank = column[:column.find('Продаж в кредит') - 1]
#         elif column.endswith('Доля кредитов'):
#             new_column_name = bank + ' ' + df_input.loc[0, column][:-1]
#             df_input.rename(columns={column: new_column_name}, inplace=True)
#    ----------------------------------------------------------
    df_input.fillna('None', inplace=True)
#    ----------------------------------------------------------
    channel = "РБТ"
    last_direction = df_input.loc[1, 'Дирекция']
    last_town = df_input.loc[1, 'Город']
    data = []
    for i in range(1, df_input.shape[0]):
        if df_input.loc[i, columns_list[0]] == 0:
                continue
        if df_input.loc[i, 'Дирекция'] == 'Местонахождение':
            break
        if df_input.loc[i, 'Дирекция'] != 'None':
                last_direction = df_input.loc[i, 'Дирекция']
        if df_input.loc[i, 'Город'] != 'None':
                last_town = df_input.loc[i, 'Город']
        need_data = []
        need_data.append(df_input.loc[i, 'Магазин'])
        need_data.append(df_input.loc[i, columns_list[0]])
        need_data.append(df_input.loc[i, columns_list[2]])
        need_data.append(channel)
        need_data.append('p')
        need_data.append(period_start)
        need_data.append(period_end)
        need_data.append(datetime.date.today())
        data.append(need_data)
#    ----------------------------------------------------------
    new_to_excel_DF = pd.DataFrame(data=data, columns=['sellerplace_group', 'sales_amt', 'sh_ko', 'channel', 'd_type', 'date_','date_to', 'date_update'])
    new_to_excel_DF.to_excel('new_dataForm_report1_period.xlsx', index=False)
    return list(new_to_excel_DF.values.tolist()), new_to_excel_DF.columns


def report2_daily(file, sheet):
    df_input = pd.read_excel(file, sheet_name=sheet, skiprows=3)  # Вставляем Excel файл с данными
    #    ----------------------------------------------------------
    columns = df_input.columns
    dates = []
    for col in columns:
        if col.startswith('Unnamed'):
            continue
        else:
            dates.append(col)
    banks = list(set(df_input.loc[0]))[1:]
    #    ----------------------------------------------------------
    for col in df_input.columns[1:]:
        if col.startswith('Unnamed'):
            df_input.rename(columns={col: last_date}, inplace=True)
        else:
            last_date = col
    df_input.rename(columns={'Unnamed: 0': 'Дирекция'}, inplace=True)
    #    ----------------------------------------------------------
    channel = "РБТ ФР"
    output_df2 = []
    last_direction = df_input.loc[1, 'Дирекция']

    for i in range(1, df_input.shape[0]):

        if df_input.loc[i, 'Дирекция'] == 'Местонахождение':
            break
        if df_input.loc[i, 'Дирекция'] != 'None':
            last_direction = df_input.loc[i, 'Дирекция']
        for j in range(1, len(df_input.columns[1:]) + 1):
            if df_input.iloc[i, j] == 0:
                continue
            output_list = []
            output_list.append('d')
            output_list.append(channel)
            output_list.append(df_input.loc[i, 'Дирекция'])
            output_list.append(datetime.datetime.strptime(df_input.columns[j].strip(), '%Y/%m/%d'))
            output_list.append(df_input.iloc[0, j])
            output_list.append(df_input.iloc[i, j])
            output_list.append(datetime.date.today())
            output_df2.append(output_list)
    #    ----------------------------------------------------------
    out_to_excel = pd.DataFrame(data=output_df2,
                                columns=['d_type', 'channel', 'sellerplace_group', 'date_', 'competitor', 'credit_amt',
                                         'date_update'])
    out_to_excel.to_excel('new_dataForm_report2_daily.xlsx', index=False)
    return list(out_to_excel.values.tolist()), out_to_excel.columns

def report2_period(file, sheet, period_start, period_end):
    df_input = pd.read_excel(file,sheet_name=sheet, skiprows=2) # Вставляем Excel файл с данными
    df_input.rename(columns={'Показатели\nas values':'Дирекция'}, inplace=True)
    columns_list = []
    columns_lower = [column.lower() for column in list(df_input.columns)]
    for column in df_input.columns[columns_lower.index('итого'):columns_lower.index('итого') + 3]:
        columns_list.append(df_input.loc[0, column])
        df_input.rename(columns={column: df_input.loc[0, column]}, inplace=True)
#    ----------------------------------------------------------
#     for bank in range(1, len(df_input.columns[1:])):
#         if df_input.columns[bank] == 'Всегда Да(ООО) ':
#             need_bank = bank
#             df_input.rename(columns={'Всегда Да(ООО) ': 'Всегда Да(ООО) Продаж в кредит'}, inplace=True)
#             break
#     df_input.rename(columns={df_input.columns[bank + 1]: 'Всегда Да(ООО) Доля кредитов'}, inplace=True)
#    ----------------------------------------------------------
    channel = "РБТ ФР"
    last_direction = df_input.loc[1, 'Дирекция']
    output_df2 = []
    for i in range(1, df_input.shape[0]):
        if df_input.loc[i, columns_list[0]] == 0:
            continue
        if df_input.loc[i, 'Дирекция'] == 'Местонахождение':
            break
        if df_input.loc[i, 'Дирекция'] != 'None':
                last_direction = df_input.loc[i, 'Дирекция']
        output_list = []
        output_list.append(last_direction)
        output_list.append(df_input.loc[i, columns_list[0]])
        output_list.append(df_input.loc[i, columns_list[2]])
        output_list.append(channel)
        output_list.append('p')
        output_list.append(period_start)
        output_list.append(period_end)
        output_list.append(datetime.date.today())
        output_df2.append(output_list)
#    ----------------------------------------------------------
    new_to_excel_DF = pd.DataFrame(data=output_df2, columns=['sellerplace_group', 'sales_amt', 'sh_ko', 'channel', 'd_type', 'date_','date_to', 'date_update'])
    new_to_excel_DF.to_excel('new_dataForm_report2_period.xlsx', index=False)
    return list(new_to_excel_DF.values.tolist()), new_to_excel_DF.columns

def actual_columns(df_input):
    k = 0
    i = 0
    for col in df_input.columns[4:]:
        if col.startswith("Unnamed"):
            k += 1
        else:
            k = 0
        if k >= 2:
            break
        i += 1
    return i + 4


columns_with_type = [('channel', 'varchar'), ('is_online', 'number'), ('sellerplace_group', 'varchar'), ('d_type', 'varchar'), ('date_', 'date'), ('date_to', 'date'), ('competitor', 'varchar'), ('credit_amt', 'number'), ('sh_ko', 'number'), ('sales_amt', 'number'), ('sh_comp', 'float'), ('date_update', 'date')]


for file in os.listdir():
    if file.endswith('xlsx'):
        if file.startswith('1'):
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb['Период']
            period_start = sheet['B2'].value
            period_end = sheet['C2'].value
            data, columns_name = report1_period(file, 'Период', period_start, period_end)
            data, columns_name = report1_daily(file, 'По дням')
        elif file.startswith('2'):
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb['Период']
            period_start = sheet['B2'].value
            period_end = sheet['C2'].value
            data, columns_name = report2_period(file, 'Период', period_start, period_end)
            data, columns_name = report2_daily(file, 'По дням')



"""
исправил:
1) wb["Период"] instead of wb.get_sheet_by_name('Период')
2)

Проблема:
Вместо даты отображаются хэш символы. Если щелкнуть на них или развдинуть ячейку,то дата появляется.
Возможные решения:
1)Увеличить размеры ячейки.
2) pd.to_exel добавить кодировку    
3) сделать чтобы формат был одинаковый. например через знаки тире
4) 
"""